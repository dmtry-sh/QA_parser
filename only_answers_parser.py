from parse_mail import *
import openpyxl
from openpyxl.styles import PatternFill


def parse(document_name, sheet_name, A_last):
    wb = openpyxl.load_workbook(filename = 'documents/{}.xlsx'.format(document_name))
    sheet = wb[sheet_name]
    
    fill_yellow = PatternFill(fill_type='solid',
                   start_color='FFFF00',
                   end_color='FFFF00')

    sheet.cell(row=1, column=2).value = 'Ответ:'

    vals = sheet['A2:A{}'.format(A_last)]
    vals = [v[0].value for v in vals]
    print('Парсер mail: запуск...')
    br = start_browser()
    i = 2
    for query in vals:
        print('Парсер mail: обрабатывается {} запрос'.format(i-1))
        try:
            answers = parse_answers_mail(br, query)
        except:
            print('Парсер mail: произошла ошибка при обработке {} запроса.'.format(i-1))
            i += 1
            continue
        for one in answers:
            j = 2
            for an in one:
                sheet.cell(row=i, column=j).fill = fill_yellow
                try:
                    sheet.cell(row=i, column=j).value = an
                    j += 1
                except:
                    sheet.cell(row=i, column=j).value = ' '
                    j += 1
        i += 1
        if (i-2) % 5 == 0:
            print('Парсер mail: сохраняю')
            wb.save('answers/{}_answers.xlsx'.format(document_name))
    print('Парсер mail: парсинг успешно завершён')
    print('Парсер mail: сохранено в файл {}_answers.xlsx'.format(document_name)) 
    wb.save('answers/{}_answers.xlsx'.format(document_name))
    close_browser(br)

if __name__ == '__main__':
    doc_name = input('Введите название документа: ')
    sheet_name = input('Введите название рабочего листа: ')
    num = int(input('Введите номер последнего запроса из таблицы: '))

    parse(doc_name, sheet_name, num)
    sleep = input('Для закрытия нажмите любую клавишу')

